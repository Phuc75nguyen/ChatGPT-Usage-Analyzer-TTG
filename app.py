import io
import json
import math
import zipfile
from datetime import datetime, timezone
from dateutil import tz
from collections import defaultdict

import pandas as pd
import streamlit as st

# add openpyxl for working with the provided Excel template
from openpyxl import load_workbook
from copy import copy

# ====== CONFIG ======
st.set_page_config(page_title="AI analysis TTG", layout="wide")
LOCAL_TZ = tz.gettz("Asia/Ho_Chi_Minh")

# ====== UTILS ======
def to_dt(ts):
    """Robust timestamp to datetime (local timezone)."""
    if ts is None or (isinstance(ts, float) and math.isnan(ts)):
        return None
    try:
        # ts is usually in seconds; sometimes string
        ts = float(ts)
        # If someone accidentally passes ms, heuristically downscale
        if ts > 1e12:  # ms
            ts = ts / 1000.0
        return datetime.fromtimestamp(ts, tz=timezone.utc).astimezone(LOCAL_TZ)
    except Exception:
        return None

def safe_get_text(msg):
    """Extract text content from ChatGPT export message blocks."""
    if not msg:
        return ""
    # v1: content could be {"content_type":"text","parts":[...]} (old)
    c = msg.get("content")
    if isinstance(c, dict):
        parts = c.get("parts")
        if isinstance(parts, list):
            return "\n".join([str(p) for p in parts if p is not None])
        # newer: content is {"content_type":"text","text": "..."} or tool calls
        if "text" in c and isinstance(c["text"], str):
            return c["text"]
    # v2: some exports put text directly at msg["content"] (string)
    if isinstance(c, str):
        return c
    return ""

def parse_conversation_json(blob, account_label=None):
    """
    Accepts either a dict loaded from conversations.json or a single conversation dict.
    Returns a flat dataframe with columns:
    account, conversation_id, conversation_title, message_id, role, author, text, create_time
    """
    rows = []
    data = blob

    # Some exports wrap in {"conversations": [...]}
    conversations = None
    if isinstance(data, dict) and "conversations" in data and isinstance(data["conversations"], list):
        conversations = data["conversations"]
    else:
        # Might be a single conversation or a top-level dict with many fields
        if isinstance(data, list):
            conversations = data
        else:
            # single conversation?
            conversations = [data]

    for conv in conversations or []:
        conv_id = conv.get("id") or conv.get("conversation_id")
        title = conv.get("title") or conv.get("conversation_title") or ""

        # Newer export: conv["mapping"] (graph of messages)
        if isinstance(conv.get("mapping"), dict):
            for node_id, node in conv["mapping"].items():
                msg = node.get("message") or {}
                if not msg:
                    continue
                role = (msg.get("author") or {}).get("role")
                author_name = (msg.get("author") or {}).get("name")
                text = safe_get_text(msg)
                ts = msg.get("create_time")
                dt = to_dt(ts)

                if role and text:
                    rows.append({
                        "account": account_label,
                        "conversation_id": conv_id,
                        "conversation_title": title,
                        "message_id": msg.get("id") or node_id,
                        "role": role,
                        "author": author_name,
                        "text": text,
                        "create_time": dt
                    })

        # Newer-ish export style: conv["messages"] = [...]
        elif isinstance(conv.get("messages"), list):
            for m in conv["messages"]:
                role = (m.get("author") or {}).get("role") or m.get("role")
                author_name = (m.get("author") or {}).get("name")
                text = safe_get_text(m)
                ts = m.get("create_time") or m.get("timestamp")  # sometimes "timestamp"
                dt = to_dt(ts)
                rows.append({
                    "account": account_label,
                    "conversation_id": conv_id,
                    "conversation_title": title,
                    "message_id": m.get("id"),
                    "role": role,
                    "author": author_name,
                    "text": text,
                    "create_time": dt
                })
        else:
            # Fallback: try best-effort common fields
            pass

    df = pd.DataFrame(rows)
    if not df.empty:
        df["date"] = df["create_time"].dt.date
        df["month"] = df["create_time"].dt.to_period("M").astype(str)
    return df

def month_range(year: int, month: int):
    start = datetime(year, month, 1, tzinfo=LOCAL_TZ)
    if month == 12:
        end = datetime(year + 1, 1, 1, tzinfo=LOCAL_TZ)
    else:
        end = datetime(year, month + 1, 1, tzinfo=LOCAL_TZ)
    return start, end

def filter_month(df, year, month):
    if df.empty:
        return df
    start, end = month_range(year, month)
    return df[(df["create_time"] >= start) & (df["create_time"] < end)].copy()

# ====== PROMPT QUALITY HEURISTICS ======
def score_prompt_quality(text: str):
    """
    Lightweight heuristic: 0-100.
    Criteria:
      - length (tokens ~ words): too short -> low; concise but structured -> good
      - has context markers (background, constraints, examples, role)
      - clear ask verbs (write, summarize, generate, translate, explain, list, compare…)
      - has evaluation/acceptance criteria (format, length caps, style)
    """
    if not text:
        return 0
    t = text.strip()
    words = len(t.split())
    score = 0

    # length
    if words < 5:
        score += 5
    elif words < 15:
        score += 20
    elif words < 60:
        score += 35
    elif words < 200:
        score += 30
    else:
        score += 25  # very long can be noisy, but still ok

    # context signals
    keywords_context = ["bối cảnh", "context", "dữ liệu", "input", "tham số", "giả định", "assume", "role:", "giả sử","bạn là"]
    if any(k.lower() in t.lower() for k in keywords_context):
        score += 15

    # constraints / format
    keywords_constraints = ["định dạng", "format", "tiêu đề", "mục", "bullet", "json", "bảng", "số từ", "ký tự", "tone"]
    if any(k.lower() in t.lower() for k in keywords_constraints):
        score += 15

    # clear ask verbs
    verbs = ["viết", "tóm tắt", "dịch", "giải thích", "liệt kê", "so sánh", "phân tích", "đưa ra", "tạo", "generate", "summarize", "translate", "explain", "viết code"]
    if any(v.lower() in t.lower() for v in verbs):
        score += 15

    # examples
    if "ví dụ" in t.lower() or "example" in t.lower() or "mẫu" in t.lower():
        score += 10

    # domain terms hint (often better prompts)
    domain_hints = ["hợp đồng", "báo cáo","phân tích","mã code", "fix code", "biên bản", "đấu thầu", "ngân sách", "marketing", "CRM", "quy trình", "SOP"]
    if any(d.lower() in t.lower() for d in domain_hints):
        score += 10

    return int(max(0, min(100, score)))

# ====== TOPIC TAGGING (EDITABLE RULES) ======
DEFAULT_TOPIC_RULES = {
    "Kinh doanh": ["doanh thu", "doanh số", "P&L", "chốt deal", "CRM", "báo cáo kinh doanh"," nhà cung cấp"],
    "Nội dung/Kế hoạch": ["quảng cáo", "chiến dịch", "meta ads", "google ads", "seo", "content", "nội dung", "thương hiệu", "branding", "viết", "văn bản"],
    "Tuyển dụng/HR": ["jd", "tuyển", "phỏng vấn", "chế độ", "lương", "onboarding", "OKR", "đánh giá", "performance", "kpi", "phúc lợi", "benefits", "hợp đồng lao động", "hợp đồng thử việc", "hợp đồng chính thức", "hợp đồng lao động thời vụ", "hợp đồng cộng tác viên", "hợp đồng khoán việc","ứng viên", "candidate", "recruitment", "hồ sơ ứng viên", "job description", "employee handbook"],
    "Pháp lý": ["hợp đồng","nhà thầu","chủ đầu tư","gói thầu", "bên a", "bên b" ,"phụ lục", "điều khoản", "pháp lý", "luật", "tuân thủ", "compliance", "kiện","tranh chấp", "thỏa thuận", "agreement", "legal","khoản","điều","nghị định", "nghị quyết", ],
    "Tài chính/Kế toán": ["hóa đơn", "ngân sách", "bút toán", "báo cáo tài chính", "thuế", "vat","invoice", "budget", "accounting", "financial report", "tax", "kiểm toán", "audit", "financial statement", "balance sheet", "profit and loss", "cash flow"],
    "Vận hành": ["quy trình", "SOP", "quy định", "quản trị", "biểu mẫu", "quản lý xây dựng"],
    "Học thuật": ["học thuật","so sánh" ,"check lại","nghiên cứu","bài báo", "tài liệu học thuật", "academic", "research", "paper","xử lý", "thế nào","lesson plan", "công thức", "cách tính"],
    "Tra cứu": ["trích","tra cứu","viết tắt","bao nhiêu","vậy còn","làm sao để " ,"là gì", "là sao", "tại sao" ,"đọc file","tìm kiếm", "lookup", "thông tin", "data", "cơ sở dữ liệu", "database", "tài liệu", "hướng dẫn", "là gì?"],
    "Dịch thuật": ["dịch","dich", "dich sang","chuyển ngữ", "Chuyển sang song ngữ Anh - Việt", "translate", "phiên dịch", "bản dịch", "ngôn ngữ", "tiếng anh", "tiếng việt"],
    "CSKH": ["hỗ trợ", "khách hàng", "phản hồi", "ticket", "trợ giúp", "giải đáp"],
    "Xây dựng": ["thi công", "công trình", "dự án", "xây dựng", "kiến trúc","bê tông", "kỹ sư", "giám sát", "thiết kế kiến trúc", "construction", "project management", "architectural design", "engineering"],
    "Báo cáo": ["báo cáo", "report", "thống kê", "phân tích", "dashboard", "biểu đồ", "chart", "data analysis"],
    "Hỗ trợ Thiết kế": ["thiết kế", "đồ họa", "sketchup", "autocad", "bản vẽ", "mockup", "prototype", "edit", "viền mỏng", "viền dày", "màu sắc", "font chữ", "logo", "branding", "design", "graphic design", "sketchup", "autocad", "photoshop", "illustrator","mm", "cm", "inch", "pixel", "dpi", "resolution", "vector", "raster", "layout", "composition", "typography"],
    "Email": ["email", "thư điện tử", "gửi mail", "trả lời email", "hộp thư", "outlook", "gmail"],
    "Kỹ thuật/Phần mềm/IT": ["api","sơ đồ","giả lập","giao diện","học sâu","học máy","chức năng","window","file","code", "router","nguồn","excel", "word", "pdf", "sharepoint", "database", "script", "server", "deploy", "bug", "log", "python", "sql","in ấn", "sửa lỗi", "sửa chữa", "bảo trì", "cài đặt", "hệ thống", "mạng", "network","AI HUb", "chatgpt", "openai", "gpt-3.5", "gpt-4", "llm", "machine learning", "deep learning", "trí tuệ nhân tạo", "artificial intelligence", "chatbot", "model", "training", "fine-tuning", "prompt engineering", "android", "ios", "app development", "web development", "frontend", "backend", "fullstack", "javascript", "react", "vue", "angular", "nodejs", "express", "django", "flask", "ruby on rails", "kích thước", "file excel"],
}

def tag_topic(text, rules):
    t = text.lower()
    for topic, kws in rules.items():
        if any(kw.lower() in t for kw in kws):
            return topic
    return "Khác"

# ====== PURPOSE TAGGING (additional heuristics for report template) ======
# The template requires a breakdown of "Mục đích sử dụng" (purpose of use) such as
# Tra cứu, tóm tắt, viết mail, học tập…  Because the raw data does not
# explicitly encode purpose, we define a simple heuristic mapping from
# keywords to broad purpose categories.  You can extend or adjust this
# dictionary to better fit your data.  Each entry maps a purpose name to a
# list of keywords; if any keyword appears in the prompt text, that purpose
# will be assigned.  If no keywords match, the prompt is tagged as "Khác".
PURPOSE_RULES = {
    "Tra cứu": [
        "tra cứu", "tra cuu", "lookup", "search", "tìm", "tìm kiếm", "là gì", "bao nhiêu", "tại sao", "how", "what",
    ],
    "Tóm tắt": [
        "tóm tắt", "tóm tat", "summary", "summarize", "tổng kết", "phân tách", "đúc kết", "tóm lược",
    ],
    "Viết mail": [
        "email", "mail", "thư", "gửi mail", "viết mail", "trả lời email", "reply email", "forward",
    ],
    "Học tập": [
        "học", "học tập", "learn", "study", "bài học", "lesson", "giảng", "nghiên cứu",
    ],
    "Dịch thuật": [
        "dịch", "dich", "translate", "dịch sang", "chuyển ngữ", "phiên dịch",
    ],
    "Viết code": [
        "code", "python", "java", "c++", "javascript", "program", "script", "sql", "lập trình",
    ],
    "Báo cáo": [
        "báo cáo", "report", "thống kê", "dashboard", "biểu đồ", "chart",
    ],
    "Kế hoạch": [
        "kế hoạch", "plan", "planning", "project", "dự án", "chiến dịch",
    ],
}

def tag_purpose(text: str, rules: dict = PURPOSE_RULES) -> str:
    """
    Assign a high‑level purpose category to a user prompt based on keyword
    matching.  Returns the first matching purpose name, or "Khác" if none
    match.  This is a simple heuristic and can be tuned by editing
    PURPOSE_RULES above.
    """
    if not text:
        return "Khác"
    t = text.lower()
    for purpose, kws in rules.items():
        if any(kw.lower() in t for kw in kws):
            return purpose
    return "Khác"

# ====== OPTIONAL: CLUSTERING (if sklearn available) ======
def try_cluster(df_user_prompts, n_clusters=15):
    try:
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.cluster import KMeans
        texts = df_user_prompts["text"].fillna("").tolist()
        if len(texts) < n_clusters:
            return None
        vec = TfidfVectorizer(max_features=5000, ngram_range=(1,2))
        X = vec.fit_transform(texts)
        km = KMeans(n_clusters=n_clusters, n_init=10, random_state=42)
        labels = km.fit_predict(X)
        return labels
    except Exception:
        return None

# ====== UI ======
st.title("📊 ChatGPT Usage Analyzer for TTG ")

st.sidebar.subheader("Tải dữ liệu")
uploaded_files = st.sidebar.file_uploader(
    "Kéo thả nhiều file conversations.json (mỗi tài khoản/phòng ban một file)",
    type=["json"], accept_multiple_files=True
)

dept_map_file = st.sidebar.file_uploader(
    "Tùy chọn: CSV map tài khoản ➜ Phòng ban (cột: account, department)",
    type=["csv"]
)

col1, col2 = st.sidebar.columns(2)
with col1:
    year = st.number_input("Năm", min_value=2023, max_value=2100, value=datetime.now().year, step=1)
with col2:
    month = st.number_input("Tháng", min_value=1, max_value=12, value=datetime.now().month, step=1)

st.sidebar.markdown("---")
st.sidebar.subheader("Cài đặt phân tích")
use_clustering = st.sidebar.checkbox("Bật clustering chủ đề (nếu có scikit‑learn)", value=False)

# Editable rules
st.sidebar.markdown("**Rules phân loại chủ đề (có thể chỉnh):**")
rules_text = st.sidebar.text_area(
    "Dạng JSON (topic -> list từ khóa)",
    value=json.dumps(DEFAULT_TOPIC_RULES, ensure_ascii=False, indent=2),
    height=220
)
try:
    topic_rules = json.loads(rules_text)
except Exception:
    st.sidebar.error("Rules JSON không hợp lệ. Dùng mặc định.")
    topic_rules = DEFAULT_TOPIC_RULES

# ====== PARSE & COMBINE ======
if uploaded_files:
    frames = []
    # Each uploaded file is expected to correspond to a single ChatGPT Plus account.
    # We use the base filename (without extension) as both the account label and
    # the department name.  This allows admins to rename the JSON files to the
    # actual department names before uploading.  If a mapping CSV is provided
    # (see below), it will override this behaviour.
    for f in uploaded_files:
        try:
            content = json.load(f)
            # derive department/account name from filename (strip extension)
            base_name = f.name.rsplit(".", 1)[0]
            account_label = base_name
            # parse conversations for this account
            df = parse_conversation_json(content, account_label=account_label)
            if not df.empty:
                # assign department equal to the file name by default
                df["department"] = base_name
            frames.append(df)
        except Exception as e:
            st.error(f"Không đọc được {f.name}: {e}")

    if frames:
        all_df = pd.concat(frames, ignore_index=True)
        # Filter to selected month
        month_df = filter_month(all_df, int(year), int(month))

        # Optionally override department names using an uploaded mapping CSV.  The CSV
        # must contain two columns: account and department.  If provided, it will
        # map the automatically derived account label (from the filename) to a
        # custom department name.  Otherwise, the department column populated
        # during parsing (equal to the filename) is retained.
        if dept_map_file is not None:
            try:
                map_df = pd.read_csv(dept_map_file)
                # accept either lowercase or original case column names
                lower_cols = [c.lower() for c in map_df.columns]
                map_df.columns = lower_cols
                if not {"account", "department"}.issubset(set(lower_cols)):
                    raise ValueError("CSV cần 2 cột: account, department")
                # perform left join to override departments based on account
                month_df = month_df.merge(map_df[["account", "department"]], on="account", how="left", suffixes=("", "_override"))
                # if override department exists, use it; otherwise keep original department
                month_df["department"] = month_df["department_override"].combine_first(month_df["department"])
                month_df.drop(columns=["department_override"], inplace=True)
            except Exception as e:
                st.error(f"Không đọc được map CSV: {e}")

        # Basic derived fields
        month_df["is_user"] = month_df["role"].eq("user")
        month_df["is_assistant"] = month_df["role"].eq("assistant")

        # Count per conversation
        conv_stats = month_df.groupby(["account","department","conversation_id","conversation_title"], dropna=False).agg(
            first_time=("create_time","min"),
            last_time=("create_time","max"),
            user_prompts=("is_user","sum"),
            assistant_msgs=("is_assistant","sum"),
            total_msgs=("message_id","count"),
            active_days=("date", lambda s: s.nunique())
        ).reset_index()

        # Topic tagging, purpose tagging & prompt quality on user prompts
        user_prompts_df = month_df[month_df["is_user"]].copy()
        user_prompts_df["topic"] = user_prompts_df["text"].apply(lambda t: tag_topic(t, topic_rules))
        user_prompts_df["purpose"] = user_prompts_df["text"].apply(lambda t: tag_purpose(t))
        user_prompts_df["prompt_quality"] = user_prompts_df["text"].apply(score_prompt_quality)
        # compute prompt word count for summary metrics
        user_prompts_df["word_count"] = user_prompts_df["text"].apply(lambda t: len(str(t).split()) if pd.notna(t) else 0)

        # Try clustering
        if use_clustering:
            labels = try_cluster(user_prompts_df)
            if labels is not None:
                user_prompts_df["cluster"] = labels
            else:
                user_prompts_df["cluster"] = None
        else:
            user_prompts_df["cluster"] = None

        # ====== OVERVIEW ======
        st.header("Tổng quan")
        total_convs = conv_stats["conversation_id"].nunique()
        total_prompts = int(user_prompts_df.shape[0])
        total_msgs = int(month_df.shape[0])
        dep_count = month_df["department"].nunique() if month_df["department"].notna().any() else 0

        kpi1, kpi2, kpi3, kpi4 = st.columns(4)
        kpi1.metric("Số cuộc hội thoại", f"{total_convs:,}")
        kpi2.metric("Tổng số prompts (user)", f"{total_prompts:,}")
        kpi3.metric("Tổng số message", f"{total_msgs:,}")
        kpi4.metric("Số phòng ban", f"{dep_count if dep_count else '-'}")

        # Aggregations
        st.subheader("Theo phòng ban")
        by_dep = conv_stats.groupby("department", dropna=False).agg(
            conversations=("conversation_id","nunique"),
            user_prompts=("user_prompts","sum"),
            assistant_msgs=("assistant_msgs","sum"),
            total_msgs=("total_msgs","sum")
        ).reset_index().rename(columns={"department":"Phòng ban"})
        st.dataframe(by_dep)

        st.subheader("Chủ đề (dựa trên rules)")
        by_topic = user_prompts_df.groupby(["department","topic"], dropna=False).size().reset_index(name="prompts")
        st.dataframe(by_topic)

        st.subheader("Điểm chất lượng prompt")
        qual = user_prompts_df.groupby("department", dropna=False)["prompt_quality"].agg(["count","mean","median","min","max"]).reset_index()
        st.dataframe(qual)

        # ====== SUMMARY FOR EXCEL TEMPLATE ======
        # Compute per‑department summary metrics required by the provided Excel template.
        # These metrics include active days, conversation counts, average message counts,
        # average prompt length (words), purpose distribution and topic distribution.
        dept_summary_rows = []
        # Determine the reporting month string (e.g., "08-2025")
        report_month_str = f"{int(month):02d}-{int(year)}"
        all_departments = sorted([d for d in month_df["department"].dropna().unique()])
        for idx, dep in enumerate(all_departments, start=1):
            # Filter conversation stats and prompts for this department
            convs = conv_stats[conv_stats["department"] == dep]
            prompts = user_prompts_df[user_prompts_df["department"] == dep]
            if convs.empty:
                continue
            # Number of conversations
            conv_count = convs["conversation_id"].nunique()
            # Average messages per conversation
            avg_msgs = float(convs["total_msgs"].mean()) if not convs.empty else 0
            # Active days: unique dates in raw messages for this department
            active_days = month_df[month_df["department"] == dep]["date"].nunique()
            # Average prompt length (words)
            avg_prompt_len = float(prompts["word_count"].mean()) if not prompts.empty else 0
            # Purpose distribution
            purpose_counts = prompts["purpose"].value_counts()
            total_p = purpose_counts.sum() if not purpose_counts.empty else 0
            purpose_lines = []
            if total_p > 0:
                for purpose_name, count in purpose_counts.items():
                    pct = count / total_p
                    # round percentage to nearest integer for display (e.g. 30%)
                    purpose_lines.append(f"+ {purpose_name} ({pct:.0%})")
            purpose_str = "\n".join(purpose_lines)
            # Topic distribution
            topic_counts = prompts["topic"].value_counts()
            total_t = topic_counts.sum() if not topic_counts.empty else 0
            topic_lines = []
            if total_t > 0:
                for topic_name, count in topic_counts.items():
                    pct = count / total_t
                    topic_lines.append(f"+ {topic_name} ({pct:.0%})")
            topic_str = "\n".join(topic_lines)
            # Build summary row
            dept_summary_rows.append({
                "Tiêu đề": dep,
                "#": idx,
                "Tháng": report_month_str,
                "Active Days (số ngày có sử dụng)": active_days,
                "Số lượng hội thoại": conv_count,
                "Số lượng tin nhắn trung bình  trong mỗi hội thoại": avg_msgs,
                "Độ dài trung bình câu prompt (từ)": avg_prompt_len,
                "Mục đích sử dụng (Tra cứu, tóm tắt, Viết mail, học tập,...) kèm %": purpose_str,
                "Chủ đề (Tuyển dụng, Thuế, ...) kèm %": topic_str,
            })
        dept_summary_df = pd.DataFrame(dept_summary_rows)

        # ====== DETAILS PER DEPARTMENT ======
        # Allow the user to drill down into each department separately.  A
        # selectbox lists all departments present in the data, and selecting
        # one will display only its conversations, prompts, replies and raw
        # messages.  Additionally, a conversation selector allows viewing
        # detailed messages in a modal popup.
        st.header("Chi tiết theo phòng ban")
        departments = sorted([d for d in month_df["department"].dropna().unique()])
        if not departments:
            st.info("Không có phòng ban nào được tìm thấy trong dữ liệu.")
        else:
            selected_dep = st.selectbox(
                "Chọn phòng ban", options=departments, index=0, key="selected_dep"
            )
            # Filter dataframes for selected department
            conv_stats_dep = conv_stats[conv_stats["department"] == selected_dep]
            user_prompts_dep = user_prompts_df[user_prompts_df["department"] == selected_dep]
            assistant_dep = month_df[(month_df["department"] == selected_dep) & (month_df["is_assistant"])]
            raw_dep = month_df[month_df["department"] == selected_dep]

            # Create tabs for each detail type within the department
            tab_a, tab_b, tab_c, tab_d = st.tabs([
                "Cuộc hội thoại", "Prompts (user)", "Assistant trả lời", "Dữ liệu gốc"
            ])

            # Conversation list with optional detail view
            with tab_a:
                st.subheader(f"Danh sách hội thoại của phòng ban {selected_dep}")
                # Display conversation summary table
                st.dataframe(
                    conv_stats_dep
                    .sort_values("first_time")
                    .rename(
                        columns={
                            "conversation_id": "ID hội thoại",
                            "conversation_title": "Tiêu đề hội thoại",
                            "first_time": "Bắt đầu",
                            "last_time": "Kết thúc",
                            "user_prompts": "Số prompt",
                            "assistant_msgs": "Số trả lời",
                            "total_msgs": "Tổng tin nhắn",
                            "active_days": "Số ngày hoạt động",
                        }
                    )
                )

                # Conversation detail selector
                if not conv_stats_dep.empty:
                    conv_choices = conv_stats_dep[["conversation_id", "conversation_title"]].apply(
                        lambda r: f"{r['conversation_id']} – {r['conversation_title']}", axis=1
                    ).tolist()
                    default_idx = 0
                    selected_conv_label = st.selectbox(
                        "Chọn hội thoại để xem chi tiết", options=conv_choices, index=default_idx, key="conv_select"
                    )
                    # Extract conversation_id from the selected label (split on the first dash)
                    selected_conv_id = selected_conv_label.split("–", 1)[0].strip()
                    # Button to open modal with conversation details
                    if st.button("Xem chi tiết", key="detail_btn"):
                        # Use a modal window if available (Streamlit >=1.25), otherwise fallback to expander
                        show_in_modal = hasattr(st, "modal")
                        container_ctx = st.modal if show_in_modal else st.expander
                        title_str = f"Chi tiết hội thoại: {selected_conv_label}"
                        with container_ctx(title_str):
                            conv_msgs = raw_dep[
                                (raw_dep["conversation_id"].astype(str) == selected_conv_id)
                            ].sort_values("create_time")
                            if conv_msgs.empty:
                                st.info("Không có dữ liệu cho hội thoại này.")
                            else:
                                # Display messages sequentially
                                for _, row in conv_msgs.iterrows():
                                    ts = row["create_time"].strftime("%Y-%m-%d %H:%M") if pd.notna(row["create_time"]) else ""
                                    if row["role"] == "user":
                                        st.markdown(f"**User ({ts})**: {row['text']}")
                                    else:
                                        st.markdown(f"**Assistant ({ts})**: {row['text']}")

            with tab_b:
                st.subheader(f"Danh sách prompt của phòng ban {selected_dep}")
                show_cols = [
                    "create_time",
                    "conversation_title",
                    "text",
                    "topic",
                    "purpose",
                    "prompt_quality",
                    "word_count",
                    "cluster",
                ]
                st.dataframe(
                    user_prompts_dep[show_cols]
                    .sort_values("create_time")
                    .rename(
                        columns={
                            "create_time": "Thời gian",
                            "conversation_title": "Tiêu đề hội thoại",
                            "text": "Nội dung prompt",
                            "topic": "Chủ đề",
                            "purpose": "Mục đích",
                            "prompt_quality": "Điểm chất lượng",
                            "word_count": "Số từ",
                            "cluster": "Nhóm",
                        }
                    )
                )

            with tab_c:
                st.subheader(f"Câu trả lời của Assistant – phòng ban {selected_dep}")
                st.dataframe(
                    assistant_dep[["create_time", "conversation_title", "text"]]
                    .sort_values("create_time")
                    .rename(
                        columns={
                            "create_time": "Thời gian",
                            "conversation_title": "Tiêu đề hội thoại",
                            "text": "Nội dung"
                        }
                    )
                )

            with tab_d:
                st.subheader(f"Dữ liệu gốc – phòng ban {selected_dep}")
                st.dataframe(
                    raw_dep.sort_values("create_time")
                )

        # ====== EXPORTS ======
        st.header("Xuất báo cáo")

        # Export summary report using the provided Excel template.  The
        # template file should exist in the current directory under the
        # name "AI Usage report Template.xlsx".  The summary dataframe
        # dept_summary_df will be inserted into the sheet named "By Dept"
        # starting from the example row.  Styles from the example row are
        # copied to preserve formatting.
        try:
            template_path = "AI Usage report Template.xlsx"
            wb = load_workbook(template_path)
            ws = wb[wb.sheetnames[0]]  # assuming the first sheet is "By Dept"
            # Identify header row containing "Tiêu đề" and the example row beneath it
            header_row_idx = None
            for row in ws.iter_rows(min_row=1, max_row=10):
                first_value = row[0].value
                if isinstance(first_value, str) and "Tiêu đề" in first_value:
                    header_row_idx = row[0].row
                    break
            if header_row_idx is None:
                raise RuntimeError("Không tìm thấy hàng tiêu đề trong template")
            example_row_idx = header_row_idx + 1
            # Determine template headers by reading the header row
            template_headers = [cell.value for cell in ws[header_row_idx]]
            # Remove any existing data rows beyond the example row
            if ws.max_row > example_row_idx:
                ws.delete_rows(example_row_idx + 1, ws.max_row - example_row_idx)
            # For each summary row, append a new row with copied styles
            for i, row_data in dept_summary_df.iterrows():
                target_row_idx = example_row_idx + i - 1  # start from example row position
                if target_row_idx > ws.max_row:
                    ws.append([None] * ws.max_column)
                for col_idx, header in enumerate(template_headers, start=1):
                    cell = ws.cell(row=target_row_idx, column=col_idx)
                    template_cell = ws.cell(row=example_row_idx, column=col_idx)
                    value = row_data.get(header, "")
                    # write value
                    cell.value = value
                    # copy style
                    cell._style = copy(template_cell._style)
            # Save workbook to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            st.download_button(
                label="⬇️ Tải Excel báo cáo",
                data=buffer.getvalue(),
                file_name=f"ai_usage_report_{int(year)}_{int(month):02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as exc:
            st.error(f"Lỗi khi tạo file Excel: {exc}")

        # Optional: export user prompts by department as a ZIP of CSVs
        if not user_prompts_df.empty:
            zbuf = io.BytesIO()
            with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
                for dep, dfdep in user_prompts_df.groupby("department"):
                    depname = "Unknown" if pd.isna(dep) else str(dep)
                    zf.writestr(f"{depname}_user_prompts.csv", dfdep.to_csv(index=False))
            st.download_button(
                "⬇️ Tải ZIP CSV theo phòng ban",
                data=zbuf.getvalue(),
                file_name=f"user_prompts_by_department_{int(year)}_{int(month):02d}.zip",
                mime="application/zip",
            )

    else:
        st.info("Chưa có dữ liệu hợp lệ.")
else:
    st.info("Hãy tải lên một hoặc nhiều file conversations.json để bắt đầu.")
